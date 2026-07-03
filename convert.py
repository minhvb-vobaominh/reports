"""
Convert Base Wework tasks export → task tracker with CR Overview.

Populates both the "Task Tracker" and "CR Overview" sheets.
Auto-detects column layout from the header row (supports the 7-col, 20-col
and 36-col Weekly-Review exports, in either .xls or .xlsx format).

Usage:
    python convert_base_to_tracker.py [source.xls|.xlsx] [template.xlsx] [output.xlsx]

Defaults:
    source   = task-input.xls
    template = task-template.xlsx
    output   = task-output.xlsx
"""

import sys
import io
import os
import re
import argparse
from collections import OrderedDict
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CLI ──────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description="Convert Base Wework export → task tracker")
    p.add_argument("source", nargs="?", default="task-input.xls", help="Source .xls from Base Wework")
    p.add_argument("template", nargs="?", default="task-template.xlsx", help="Template .xlsx")
    p.add_argument("output", nargs="?", default="task-output.xlsx", help="Output .xlsx")
    return p.parse_args()


# ── Styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
BODY_FONT   = Font(name="Segoe UI", size=10, color="333333")
BOLD_FONT   = Font(name="Segoe UI", size=10, bold=True, color="333333")
MUTED_FONT  = Font(name="Segoe UI", size=9, color="64748B", italic=True)
BORDER = Border(bottom=Side(style="thin", color="D0D9E8"),
                right=Side(style="thin", color="D0D9E8"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

STATUS_FILL = {
    "Done":           PatternFill("solid", fgColor="DCFCE7"),
    "Doing":          PatternFill("solid", fgColor="DBEAFE"),
    "Overdue":        PatternFill("solid", fgColor="FEE2E2"),
    "Blocked":        PatternFill("solid", fgColor="FEF3C7"),
    "Todo":           PatternFill("solid", fgColor="F1F5F9"),
    "Not Started":    PatternFill("solid", fgColor="F1F5F9"),
    "In Progress":    PatternFill("solid", fgColor="DBEAFE"),
    "Partially Done": PatternFill("solid", fgColor="FEF9C3"),
    "Rejected":       PatternFill("solid", fgColor="E5E7EB"),
    "UAT":            PatternFill("solid", fgColor="EDE9FE"),
}
STATUS_FONT = {
    "Done":           Font(name="Segoe UI", size=10, bold=True, color="15803D"),
    "Doing":          Font(name="Segoe UI", size=10, bold=True, color="1D4ED8"),
    "Overdue":        Font(name="Segoe UI", size=10, bold=True, color="B91C1C"),
    "Blocked":        Font(name="Segoe UI", size=10, bold=True, color="92400E"),
    "Todo":           Font(name="Segoe UI", size=10, color="475569"),
    "Not Started":    Font(name="Segoe UI", size=10, color="475569"),
    "In Progress":    Font(name="Segoe UI", size=10, bold=True, color="1D4ED8"),
    "Partially Done": Font(name="Segoe UI", size=10, bold=True, color="854D0E"),
    "Rejected":       Font(name="Segoe UI", size=10, color="6B7280", strike=True),
    "UAT":            Font(name="Segoe UI", size=10, bold=True, color="6D28D9"),
    "\u2014":         Font(name="Segoe UI", size=10, color="CBD5E1"),
}
PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="FEE2E2"),
    "High":     PatternFill("solid", fgColor="FEF3C7"),
    "Medium":   PatternFill("solid", fgColor="DBEAFE"),
    "Low":      PatternFill("solid", fgColor="F1F5F9"),
}


# ── Team lead mapping ───────────────────────────────────────────────
LEAD_MAP = {
    "phongvo2440":    ("Team Mobile", "phongvo@phs.vn"),
    "minhvo2565":     ("Dev Team B",  "minhvo@phs.vn"),
    "huunguyen2525":  ("MW Team",     "huunguyen@phs.vn"),
}

# "Project Owner" column values → canonical team names
PROJECT_OWNER_MAP = {
    "Team B":      "Dev Team B",
    "Dev Team B":  "Dev Team B",
    "Mobile Team": "Team Mobile",
    "Team Mobile": "Team Mobile",
    "MW Team":     "MW Team",
}

TEAM_LEADS = {team: email for team, email in LEAD_MAP.values()}


# ── Status mapping (Vietnamese → English) ───────────────────────────
STATUS_MAP = {
    "Hoàn thành":      "Done",
    "Hoàn thành muộn": "Done",   # completed late → still Done
    "Đang làm":        "Doing",
    "Quá hạn":         "Overdue",
    "Chưa bắt đầu":   "Todo",
}

# ── Stage mapping (CR lifecycle → status) ───────────────────────────
# The Stage column reflects where the CR actually is in its lifecycle;
# it takes precedence over the WeWork task status (a task can be marked
# "Hoàn thành" while the CR is still in UAT).
STAGE_STATUS_MAP = {
    "Go-live":           "Done",
    "Ready for go-live": "Done",
    "Rejected":          "Rejected",
    "UAT Testing":       "UAT",
    "FTL Coding":        "Doing",
    "Web B Coding":      "Doing",
    "Design UI/UX":      "Doing",
    "Design review":     "Doing",
    "FSS Eval CR":       "Doing",
    "Discuss Req":       "Doing",
    "Waiting FTL":       "Blocked",
    "Pending":           "Blocked",
    "Backlog":           "Todo",
    "Dev To do":         "Todo",
}

# Keyword fallback for stage values not in the exact map
STAGE_KEYWORD_RULES = [
    (("go-live", "golive", "done"),                          "Done"),
    (("reject", "cancel"),                                    "Rejected"),
    (("waiting", "pending", "hold"),                          "Blocked"),
    (("backlog", "to do", "todo"),                            "Todo"),
    (("uat", "sit", "testing"),                               "UAT"),
    (("coding", "design", "review", "eval",
      "discuss", "req", "dev"),                               "Doing"),
]


# Lifecycle order for funnel / CFD sorting (−1 = out of flow)
STAGE_ORDER = {
    "Backlog": 0,
    "Discuss Req": 1, "FSS Eval CR": 1,
    "Design UI/UX": 2, "Design review": 2,
    "Dev To do": 3, "FTL Coding": 3, "Web B Coding": 3,
    "Waiting FTL": 3, "Pending": 3,
    "UAT Testing": 5,
    "Ready for go-live": 6,
    "Go-live": 7,
    "Rejected": -1,
}

DATE_FORMATS = ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d")


def to_datetime(value, datemode=None):
    """Parse any raw cell into a datetime, or None.
    Handles datetime objects (openpyxl), float serials (xlrd) and strings."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, float) and value > 0 and datemode is not None:
        try:
            return xlrd.xldate_as_datetime(value, datemode)
        except Exception:
            return None
    if isinstance(value, str) and value.strip():
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def split_multi(value):
    """Split a multi-valued cell like 'RMD, IT' into a clean list."""
    if not value:
        return []
    return [p.strip() for p in re.split(r"[,;]", str(value)) if p.strip()]


def derive_risk(t, report_date):
    """Risk heuristic for the fact table."""
    if t["status"] in ("Overdue", "Blocked"):
        return "High"
    active = t["status"] in ("Doing", "UAT")
    if active and t["deadline_at"] and t["deadline_at"].date() < report_date:
        return "High"
    if active and not t["deadline_at"]:
        return "Medium"
    if active and t["stale_days"] is not None and t["stale_days"] > 14:
        return "Medium"
    return "Low"


def read_team_config(wb):
    """Read team → lead email from the template's Team Config sheet.
    Returns (team_leads dict, teams list); falls back to code defaults."""
    if "Team Config" not in wb.sheetnames:
        return dict(TEAM_LEADS), list(TEAM_LEADS)
    leads, teams = {}, []
    for row in wb["Team Config"].iter_rows(min_row=2, values_only=True):
        name, _lead, email = (row + (None,) * 3)[:3]
        if name and email:
            leads[str(name).strip()] = str(email).strip()
            teams.append(str(name).strip())
    return (leads, teams) if leads else (dict(TEAM_LEADS), list(TEAM_LEADS))


def stage_to_status(stage):
    """Map a Stage value to a status, or None if the stage is unknown."""
    stage = stage.strip()
    if not stage:
        return None
    if stage in STAGE_STATUS_MAP:
        return STAGE_STATUS_MAP[stage]
    low = stage.lower()
    for keywords, status in STAGE_KEYWORD_RULES:
        if any(k in low for k in keywords):
            return status
    return None


def derive_status(stage, status_vn):
    """Decide task status: Stage drives it, WeWork status is the fallback,
    and an overdue WeWork status overrides any not-yet-done stage."""
    wework = STATUS_MAP.get(status_vn, status_vn)
    staged = stage_to_status(stage)
    if staged is None:
        return wework
    if wework == "Overdue" and staged not in ("Done", "Rejected"):
        return "Overdue"
    return staged

# ── Base Wework column indices ──────────────────────────────────────
# Auto-detected from header row; these are the defaults for the 20-col export.
COL_DEFAULTS = {
    "task_name":    0,   # Tên công việc
    "assigner":     1,   # Người giao việc
    "executor":     2,   # Người thực hiện
    "watcher":      3,   # Người theo dõi
    "urgent":       4,   # Khẩn cấp
    "important":    5,   # Quan trọng
    "labels":       6,   # Danh sách nhãn
    "start_date":   7,   # Ngày bắt đầu
    "deadline":     8,   # Thời hạn
    "completed":    9,   # Hoàn thành thực tế
    "description":  10,  # Mô tả công việc
    "status":       11,  # Trạng thái
    "result":       12,  # Kết quả công việc
    "objective":    13,  # Mục tiêu
    "created":      14,  # Ngày tạo
    "task_id":      15,  # Mã công việc (ID)
    "parent_id":    16,  # Mã công việc cha (ID)
    "metatype":     17,  # Metatype
    "executor_name": 18, # Họ và tên người nhận việc
    "assigner_name": 19, # Họ và tên người giao việc
}

# Known header names → column keys (Vietnamese)
HEADER_MAP = {
    "Tên công việc":              "task_name",
    "Người giao việc":            "assigner",
    "Người thực hiện":            "executor",
    "Người theo dõi":             "watcher",
    "Khẩn cấp":                   "urgent",
    "Quan trọng":                  "important",
    "Danh sách nhãn":             "labels",
    "Ngày bắt đầu":              "start_date",
    "Thời hạn":                    "deadline",
    "Hoàn thành thực tế":         "completed",
    "Mô tả công việc":           "description",
    "Trạng thái":                  "status",
    "Kết quả công việc":          "result",
    "Mục tiêu":                    "objective",
    "Ngày tạo":                    "created",
    "Mã công việc (ID)":          "task_id",
    "Mã công việc cha (ID)":      "parent_id",
    "Metatype":                    "metatype",
    "Họ và tên người nhận việc":  "executor_name",
    "Họ và tên người giao việc":  "assigner_name",
    "Project Owner":               "project_owner",
    "Stage":                       "stage",
    "Owner division":              "owner_division",
    "Platform":                    "platform",
    "Related divisions":           "related_divisions",
    "Q-KPI":                       "q_kpi",
    "Detail":                      "detail",
    "CR 1":                        "cr1",
    "CR 2":                        "cr2",
    "Dev in charge":               "dev_in_charge",
    "By 30/03":                    "by_3003",
    "Reason / Condition":          "reason",
    "Weekly-Review - Assessment":  "wr_assessment",
    "Weekly-Review - Next Actions": "wr_next",
    "Weekly-Review - Date Review": "wr_review_date",
}


def detect_columns(header_row):
    """Auto-detect column indices from the header row.
    Falls back to COL_DEFAULTS if headers don't match."""
    cols = dict(COL_DEFAULTS)
    matched = 0
    for col_idx, header in enumerate(header_row):
        if str(header).strip() in HEADER_MAP:
            cols[HEADER_MAP[str(header).strip()]] = col_idx
            matched += 1
    return cols, matched


def read_source(path):
    """Read the source export into a plain grid of raw cell values.
    Supports .xls (xlrd) and .xlsx (openpyxl).
    Returns (rows, datemode) — datemode is only meaningful for .xls."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        # openpyxl may return ragged rows in read-only mode — pad to header width
        width = max((len(r) for r in rows), default=0)
        for r in rows:
            r.extend([None] * (width - len(r)))
        return rows, None
    src_wb = xlrd.open_workbook(path)
    src_ws = src_wb.sheet_by_index(0)
    rows = [[src_ws.cell_value(r, c) for c in range(src_ws.ncols)]
            for r in range(src_ws.nrows)]
    return rows, src_wb.datemode


# ── Helpers ──────────────────────────────────────────────────────────
def parse_request_id(task_name):
    """Extract request-id from task name like [CR] [2275303] ... or [SDK] [2400512] ...
    Also handles ranges like [CR] [2310394 - 2391899] (uses the first id)."""
    m = re.search(r"\[(?:CR|SDK|INB)\]\s*\[(\d+)(?:\s*-\s*\d+)?\]", task_name)
    if m:
        return m.group(1)
    return None


def parse_dept_tag(task_name):
    """Extract department tag like [RMD], [SS], [MKT] from task name."""
    # Skip the first bracket group ([CR]/[SDK]) and request-id, find dept tag
    m = re.search(r"\[\d+(?:\s*-\s*\d+)?\]\s*\[([A-Z]{2,4})\]", task_name)
    if m:
        return m.group(1)
    return ""


def parse_category(task_name):
    """Determine category from task name prefix."""
    if task_name.startswith("[SDK]"):
        return "SDK"
    if task_name.startswith("[CR]") or task_name.startswith("[CR]["):
        return "CR"
    if task_name.startswith("[MKT]") or task_name.startswith("[INB]"):
        return "Internal"
    return "CR"


def parse_date_cell(value, datemode):
    """Parse a date cell — handles strings, datetime objects (openpyxl)
    and float serial dates (xlrd)."""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, float) and value > 0 and datemode is not None:
        try:
            dt = xlrd.xldate_as_datetime(value, datemode)
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return ""
    if isinstance(value, str) and value.strip():
        return value.strip()
    return ""


def clean_title(task_name):
    """Extract a clean CR title from the full task name.
    Removes [CR], [request-id], [dept], [No.XX-YYYY] prefixes."""
    title = task_name
    # Remove leading [CR] / [SDK] / [INB]
    title = re.sub(r"^\[(?:CR|SDK|INB)\]\s*", "", title)
    # Remove [request-id] (single id or range like [2310394 - 2391899])
    title = re.sub(r"^\[?\d+(?:\s*-\s*\d+)?\]?\s*", "", title)
    # Remove [dept]
    title = re.sub(r"^\[[A-Z]{2,4}\]\s*", "", title)
    # Remove [No.XX-YYYY]
    title = re.sub(r"^\[No\.\d+-\d+\]\s*", "", title)
    return title.strip()


def detect_section(section_name):
    """Parse section header to get team and phase.
    e.g. 'Team Mobile - Q2:' → ('Team Mobile', 'Q2')
         'Dev Team B - Backlog:' → ('Dev Team B', 'Backlog')
         'CRs in Coding:' → (None, 'Q2')  # cross-team, use lead to determine
    """
    section_name = section_name.rstrip(":")
    if "CRs in Coding" in section_name:
        return None, "Q2"

    for team in ["Team Mobile", "Dev Team B", "MW Team"]:
        if team in section_name:
            if "Backlog" in section_name:
                return team, "Backlog"
            m = re.search(r"Q(\d)", section_name)
            if m:
                return team, f"Q{m.group(1)}"
            return team, "Q2"
    return None, "Q2"


def derive_overall_status(statuses):
    """Derive overall CR status from per-team statuses."""
    present = [s for s in statuses if s != "\u2014"]
    # Rejected teams don't count toward progress; all-rejected CR is Rejected
    active = [s for s in present if s != "Rejected"]
    if present and not active:
        return "Rejected"
    if not active:
        return "Not Started"
    if any(s == "Overdue" for s in active):
        return "Overdue"
    if any(s == "Blocked" for s in active):
        return "Blocked"
    if all(s == "Done" for s in active):
        return "Done"
    if all(s == "UAT" for s in active):
        return "UAT"
    if all(s == "Todo" for s in active):
        return "Not Started"
    if any(s == "Done" for s in active) and any(s != "Done" for s in active):
        return "Partially Done"
    return "In Progress"


def derive_priority(statuses_with_deadlines):
    """Simple priority heuristic based on overdue count and deadline proximity."""
    overdue_count = sum(1 for s, _ in statuses_with_deadlines if s == "Overdue")
    if overdue_count >= 2:
        return "Critical"
    if overdue_count == 1:
        return "High"
    has_deadline = any(d for _, d in statuses_with_deadlines)
    if has_deadline:
        return "Medium"
    return "Low"


# ── Star-schema export (fact / dim / bridge sheets + snapshot CSV) ──
FACT_FIELDS = [
    "task_id", "request_id", "team", "assignee", "stage", "stage_order",
    "status", "priority", "phase", "q_kpi", "is_subtask", "is_done",
    "wip_flag", "is_multi_team", "team_count", "has_deadline", "sla_met",
    "completed_late", "late_days", "days_to_deadline", "age_days",
    "cycle_time_days", "lead_time_days", "stale_days", "risk_level",
    "waiting_external", "has_weekly_review", "recommit_count",
]

DIM_TASK_FIELDS = [
    "task_id", "request_id", "task_name", "category", "dept_tag",
    "metatype", "parent_id", "jira_url", "owner_division",
    "created_at", "start_at", "deadline_at", "completed_at",
    "cr_received", "cr_recommit", "last_review",
    "reason", "wr_assessment", "wr_next",
]


def _write_table(wb, name, headers, rows):
    """(Re)create a flat data sheet with a styled header row."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"
    ws.freeze_panes = "A2"


def write_star_schema(wb, tasks, team_leads, report_date):
    """Write fact/dim/bridge sheets for BI tools (Power BI, Metabase, ...)."""
    fact_rows = [[report_date.isoformat()] + [t[f] for f in FACT_FIELDS]
                 for t in tasks]
    _write_table(wb, "fact_task_snapshot",
                 ["report_date"] + FACT_FIELDS, fact_rows)

    _write_table(wb, "dim_task", DIM_TASK_FIELDS,
                 [[t[f] for f in DIM_TASK_FIELDS] for t in tasks])

    _write_table(wb, "dim_team", ["team", "lead_email"],
                 [[team, email] for team, email in team_leads.items()])

    _write_table(wb, "dim_stage", ["stage", "stage_order", "status_group"],
                 [[s, o, stage_to_status(s)] for s, o in
                  sorted(STAGE_ORDER.items(), key=lambda x: x[1])])

    for name, field in [("bridge_task_division", "related_divisions"),
                        ("bridge_task_platform", "platform"),
                        ("bridge_task_devunit", "dev_in_charge")]:
        rows = [[t["task_id"], v] for t in tasks for v in split_multi(t[field])]
        _write_table(wb, name, ["task_id", name.split("_")[-1]], rows)


def append_snapshot_csv(tasks, report_date, output_path):
    """Append today's snapshot to fact_task_snapshot.csv next to the output.
    Re-running on the same day replaces that day's rows (idempotent)."""
    import csv
    csv_path = os.path.join(os.path.dirname(output_path) or ".",
                            "fact_task_snapshot.csv")
    header = ["report_date"] + FACT_FIELDS
    old_rows = []
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)
            old_rows = [r for r in reader if r and r[0] != report_date.isoformat()]

    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, datetime):
            return v.isoformat(sep=" ")
        return str(v)

    new_rows = [[report_date.isoformat()] + [fmt(t[f]) for f in FACT_FIELDS]
                for t in tasks]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(old_rows)
        writer.writerows(new_rows)
    return csv_path, len(old_rows), len(new_rows)


def write_dashboard_js(tasks, team_leads, report_date, output_path):
    """Emit dashboard-data.js next to the output so index.html can load
    task-level data via <script src> (works over file:// with no server)."""
    import json
    js_path = os.path.join(os.path.dirname(output_path) or ".", "dashboard-data.js")

    def iso(dt):
        return dt.strftime("%Y-%m-%d") if dt else None

    recs = []
    for t in tasks:
        recs.append({
            "task_id":     t["task_id"],
            "request_id":  t["request_id"],
            "title":       clean_title(t["task_name"]),
            "team":        t["team"],
            "category":    t["category"],
            "dept":        t["dept_tag"],
            "priority":    t["priority"],
            "status":      t["status"],
            "stage":       t["stage"],
            "stage_order": t["stage_order"],
            "phase":       t["phase"],
            "q_kpi":       t["q_kpi"],
            "assignee":    t["assignee"],
            "deadline":    iso(t["deadline_at"]),
            "completed":   iso(t["completed_at"]),
            "late_days":   t["late_days"],
            "sla_met":     t["sla_met"],
            "completed_late": t["completed_late"],
            "risk":        t["risk_level"],
            "multi_team":  t["is_multi_team"],
            "team_count":  t["team_count"],
            "subtask":     t["is_subtask"],
            "jira":        t["jira_url"],
            "wr_assessment": t["wr_assessment"],
            "wr_next":     t["wr_next"],
            "wr_date":     iso(t["last_review"]),
        })

    payload = (
        "// Auto-generated by convert_base_to_tracker.py — do not edit by hand.\n"
        f"const REPORT_DATE = {json.dumps(report_date.isoformat())};\n"
        f"const TEAMS = {json.dumps(list(team_leads), ensure_ascii=False)};\n"
        f"const FACT = {json.dumps(recs, ensure_ascii=False, indent=1)};\n"
    )
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(payload)
    return js_path


# ── Main ─────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    source_path   = os.path.join(script_dir, args.source)   if not os.path.isabs(args.source)   else args.source
    template_path = os.path.join(script_dir, args.template)  if not os.path.isabs(args.template) else args.template
    output_path   = os.path.join(script_dir, args.output)    if not os.path.isabs(args.output)   else args.output

    # ── Read source (.xls or .xlsx) ──────────────────────────────────
    grid, datemode = read_source(source_path)
    ncols = len(grid[0]) if grid else 0

    print(f"Source: {source_path} ({len(grid)} rows, {ncols} cols)")

    # Auto-detect column layout from header row
    C, matched = detect_columns(grid[0] if grid else [])
    print(f"Column detection: {matched}/{len(HEADER_MAP)} headers matched")

    # ── Load template early: Team Config drives team/lead resolution ─
    wb = load_workbook(template_path)
    team_leads, teams_list = read_team_config(wb)
    TEAM_LEADS.update(team_leads)
    report_date = datetime.now().date()

    # Parse all task rows, skipping headers and section dividers
    tasks = []
    current_section_team = None
    current_phase = "Q2"

    def cell_raw(row, key):
        col = C.get(key)
        if col is None or col >= len(grid[row]):
            return None
        return grid[row][col]

    def cell_str(row, key):
        v = cell_raw(row, key)
        return "" if v is None else str(v).strip()

    for row_idx in range(1, len(grid)):
        task_name  = cell_str(row_idx, "task_name")
        assigner   = cell_str(row_idx, "assigner")
        deadline   = cell_raw(row_idx, "deadline")
        completed  = cell_raw(row_idx, "completed")
        status_vn  = cell_str(row_idx, "status")
        created    = cell_str(row_idx, "created")
        task_id    = cell_str(row_idx, "task_id")
        executor   = cell_str(row_idx, "executor") if "executor" in C else ""
        executor_name = cell_str(row_idx, "executor_name") if "executor_name" in C else ""
        project_owner = cell_str(row_idx, "project_owner") if "project_owner" in C else ""
        parent_id  = cell_str(row_idx, "parent_id") if "parent_id" in C else ""
        stage      = cell_str(row_idx, "stage") if "stage" in C else ""
        start_raw  = cell_raw(row_idx, "start_date")
        metatype   = cell_str(row_idx, "metatype") if "metatype" in C else ""
        owner_division = cell_str(row_idx, "owner_division") if "owner_division" in C else ""
        platform   = cell_str(row_idx, "platform") if "platform" in C else ""
        related_divisions = cell_str(row_idx, "related_divisions") if "related_divisions" in C else ""
        q_kpi      = cell_str(row_idx, "q_kpi") if "q_kpi" in C else ""
        jira_url   = cell_str(row_idx, "detail") if "detail" in C else ""
        cr1_raw    = cell_raw(row_idx, "cr1")
        cr2_raw    = cell_raw(row_idx, "cr2")
        dev_in_charge = cell_str(row_idx, "dev_in_charge") if "dev_in_charge" in C else ""
        reason     = cell_str(row_idx, "reason") if "reason" in C else ""
        wr_assessment = cell_str(row_idx, "wr_assessment") if "wr_assessment" in C else ""
        wr_next    = cell_str(row_idx, "wr_next") if "wr_next" in C else ""
        wr_review_raw = cell_raw(row_idx, "wr_review_date")

        # Section header row (no assigner, name ends with ':')
        if not assigner and ":" in task_name:
            current_section_team, current_phase = detect_section(task_name)
            continue

        # Skip empty rows
        if not task_name:
            continue

        # Determine team: section header → assigner lead map → Project Owner column
        if current_section_team:
            team = current_section_team
        elif assigner in LEAD_MAP:
            team = LEAD_MAP[assigner][0]
        elif project_owner in PROJECT_OWNER_MAP:
            team = PROJECT_OWNER_MAP[project_owner]
        else:
            team = "Unknown"

        # Get lead email: assigner lead map, else the resolved team's lead
        if assigner in LEAD_MAP:
            lead_email = LEAD_MAP[assigner][1]
        else:
            lead_email = TEAM_LEADS.get(team, assigner)

        # Parse fields
        request_id  = parse_request_id(task_name)
        dept_tag    = parse_dept_tag(task_name)
        category    = parse_category(task_name)
        status      = derive_status(stage, status_vn)
        created_at    = to_datetime(cell_raw(row_idx, "created"), datemode)
        start_at      = to_datetime(start_raw, datemode)
        deadline_at   = to_datetime(deadline, datemode)
        completed_at  = to_datetime(completed, datemode)
        cr_received   = to_datetime(cr1_raw, datemode)
        cr_recommit   = to_datetime(cr2_raw, datemode)
        last_review   = to_datetime(wr_review_raw, datemode)

        # Backlog tasks → Todo
        phase = current_phase
        if phase == "Backlog" and status == "Doing":
            status = "Todo"

        # Clean up task_id / parent_id (remove .0 from float)
        if task_id.endswith(".0"):
            task_id = task_id[:-2]
        if parent_id.endswith(".0"):
            parent_id = parent_id[:-2]

        # Determine assignee: prefer executor_name (full name), fall back to executor (username)
        assignee = executor_name if executor_name else executor

        tasks.append({
            "request_id": request_id or "",
            "task_id":    task_id,
            "parent_id":  parent_id,
            "task_name":  task_name,
            "team":       team,
            "category":   category,
            "dept_tag":   dept_tag,
            "priority":   "",  # will be set later
            "status":     status,
            "stage":      stage,
            "phase":      phase,
            "assignee":   assignee,
            "lead_email": lead_email,
            "created":    created_at,
            "deadline":   deadline_at,
            "completed":  completed_at,
            "dependencies": "",
            "notes":      "",
            # raw datetimes (aliases used by the star-schema export)
            "created_at":   created_at,
            "start_at":     start_at,
            "deadline_at":  deadline_at,
            "completed_at": completed_at,
            "cr_received":  cr_received,
            "cr_recommit":  cr_recommit,
            "last_review":  last_review,
            # extra raw attributes
            "metatype":     metatype,
            "owner_division": owner_division,
            "platform":     platform,
            "related_divisions": related_divisions,
            "q_kpi":        q_kpi,
            "jira_url":     jira_url,
            "dev_in_charge": dev_in_charge,
            "reason":       reason,
            "wr_assessment": wr_assessment,
            "wr_next":      wr_next,
            "completed_late": status_vn == "Hoàn thành muộn",
        })

    # Subtasks with no resolvable team inherit team/lead from their parent task
    tasks_by_id = {t["task_id"]: t for t in tasks if t["task_id"]}
    for t in tasks:
        if t["team"] == "Unknown" and t["parent_id"] in tasks_by_id:
            parent = tasks_by_id[t["parent_id"]]
            if parent["team"] != "Unknown":
                t["team"] = parent["team"]
                t["lead_email"] = TEAM_LEADS.get(parent["team"], t["lead_email"])

    # ── Drop duplicated CR tasks ─────────────────────────────────────
    # The same CR is sometimes tracked twice in WeWork within one team
    # (task re-created). Keep only the most recently created task when
    # request_id + team + normalized task name match. Tasks with different
    # names (e.g. "Phase 1" vs "Phase 2" of one CR) are NOT duplicates.
    def _dup_key(t):
        title = re.sub(r"\s+", " ", t["task_name"].strip().lower())
        return (t["request_id"], t["team"], title)

    newest = {}
    for t in tasks:
        if not t["request_id"]:
            continue
        k = _dup_key(t)
        cur = newest.get(k)
        if cur is None or (t["created_at"] or datetime.min) > (cur["created_at"] or datetime.min):
            newest[k] = t

    dropped = [t for t in tasks
               if t["request_id"] and newest[_dup_key(t)] is not t]
    if dropped:
        tasks = [t for t in tasks
                 if not t["request_id"] or newest[_dup_key(t)] is t]
        for t in dropped:
            print(f"  Duplicate dropped: CR {t['request_id']} / {t['team']} — task {t['task_id']} ({t['task_name'][:50]})")
        print(f"Deduplicated: {len(dropped)} duplicate CR task(s) removed")

    print(f"Parsed: {len(tasks)} tasks")

    # ── Build CR groups ──────────────────────────────────────────────
    # Group by request_id to detect multi-team CRs and derive overview
    cr_groups = OrderedDict()  # request_id → list of task dicts
    no_id_tasks = []

    for t in tasks:
        rid = t["request_id"]
        if rid:
            cr_groups.setdefault(rid, []).append(t)
        else:
            no_id_tasks.append(t)

    # For multi-team CRs, set dependencies and priority
    for rid, group in cr_groups.items():
        teams_in_cr = set(t["team"] for t in group)
        is_multi = len(teams_in_cr) > 1

        # Priority heuristic
        statuses_deadlines = [(t["status"], t["deadline"]) for t in group]
        priority = derive_priority(statuses_deadlines)

        for t in group:
            t["priority"] = priority
            if is_multi:
                t["dependencies"] = rid  # link to same CR

    # Single-team CRs without priority yet
    for t in tasks:
        if not t["priority"]:
            if t["status"] == "Overdue":
                t["priority"] = "High"
            elif t["phase"] == "Backlog":
                t["priority"] = "Low"
            else:
                t["priority"] = "Medium"

    # ── Derived metrics for the star schema ──────────────────────────
    for t in tasks:
        rid = t["request_id"]
        group = cr_groups.get(rid, [t])
        t["team_count"] = len(set(g["team"] for g in group))
        t["is_multi_team"] = t["team_count"] > 1
        t["is_subtask"] = t["metatype"] == "Subtask"
        t["stage_order"] = STAGE_ORDER.get(t["stage"], None)
        t["is_done"] = t["status"] in ("Done", "Rejected")
        t["wip_flag"] = t["status"] in ("Doing", "UAT")
        t["has_deadline"] = t["deadline_at"] is not None
        t["has_weekly_review"] = bool(t["wr_assessment"])
        t["recommit_count"] = 1 if t["cr_recommit"] else 0
        t["waiting_external"] = t["stage"] in ("Waiting FTL", "Pending")

        d = t["deadline_at"]
        c = t["completed_at"]
        t["sla_met"] = (c <= d) if (c and d) else None
        if c and d and c > d:
            t["late_days"] = (c - d).days
        elif not c and d and d.date() < report_date and not t["is_done"]:
            t["late_days"] = (report_date - d.date()).days
        else:
            t["late_days"] = 0
        t["days_to_deadline"] = (d.date() - report_date).days if (d and not c) else None
        t["age_days"] = (report_date - t["created_at"].date()).days if t["created_at"] else None
        base = t["start_at"] or t["created_at"]
        t["cycle_time_days"] = (c - base).days if (c and base) else None
        t["lead_time_days"] = (c - t["cr_received"]).days if (c and t["cr_received"]) else None
        t["stale_days"] = (report_date - t["last_review"].date()).days if t["last_review"] else None
        t["risk_level"] = derive_risk(t, report_date)

    # ── Populate Task Tracker ────────────────────────────────────────
    ws_tracker = wb["Task Tracker"]

    # Insert a "Stage" column after Status (col 8) if the template lacks it
    if ws_tracker.cell(row=2, column=9).value != "Stage":
        # Unmerge the title row first — openpyxl doesn't shift merged
        # ranges on insert_cols — then re-merge across the widened table
        for m in [m for m in ws_tracker.merged_cells.ranges if m.min_row == 1]:
            ws_tracker.unmerge_cells(str(m))
        ws_tracker.insert_cols(9)
        ws_tracker.merge_cells("A1:Q1")
        hdr = ws_tracker.cell(row=2, column=9, value="Stage")
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        hdr.alignment = CENTER
        hdr.border = BORDER
        ws_tracker.column_dimensions["I"].width = 16

    # Clear existing sample data (rows 3+)
    for row in range(3, ws_tracker.max_row + 1):
        for col in range(1, 18):
            ws_tracker.cell(row=row, column=col).value = None
            ws_tracker.cell(row=row, column=col).font = BODY_FONT
            ws_tracker.cell(row=row, column=col).fill = PatternFill()

    # Write tasks
    tracker_fields = [
        "request_id", "task_id", "task_name", "team", "category",
        "dept_tag", "priority", "status", "stage", "phase", "assignee",
        "lead_email", "created", "deadline", "completed",
        "dependencies", "notes",
    ]
    wide_cols = {2, 15, 16}  # 0-indexed: task_name, dependencies, notes

    for i, t in enumerate(tasks, start=3):
        for j, field in enumerate(tracker_fields):
            cell = ws_tracker.cell(row=i, column=j + 1, value=t[field])
            cell.font = BODY_FONT
            cell.alignment = LEFT if j in wide_cols else CENTER
            cell.border = BORDER
            if isinstance(t[field], datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"
        ws_tracker.row_dimensions[i].height = 24

    tracker_count = len(tasks)
    print(f"Task Tracker: {tracker_count} rows written")

    # ── Populate CR Overview ─────────────────────────────────────────
    # Slim templates may only ship the Task Tracker sheet — build the
    # CR Overview sheet (title + header row) from scratch if missing.
    if "CR Overview" not in wb.sheetnames:
        ws_overview = wb.create_sheet("CR Overview")
        ws_overview.merge_cells("A1:J1")
        title = ws_overview.cell(row=1, column=1, value="CR Overview")
        title.font = Font(name="Segoe UI", size=14, bold=True, color="1B2A4A")
        headers = ["Request ID", "CR Title", "Dept", "Priority", "Teams",
                   "Team Mobile", "Dev Team B", "MW Team", "Overall Status", "Notes"]
        for ci, h in enumerate(headers, 1):
            c = ws_overview.cell(row=3, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = BORDER
        for col, w in zip("ABCDEFGHIJ", [12, 46, 8, 10, 10, 14, 14, 14, 16, 40]):
            ws_overview.column_dimensions[col].width = w
        ws_overview.freeze_panes = "A4"
    ws_overview = wb["CR Overview"]

    # Unmerge all merged cells in rows 4+ before clearing
    merges_to_remove = [m for m in ws_overview.merged_cells.ranges if m.min_row >= 4]
    for m in merges_to_remove:
        ws_overview.unmerge_cells(str(m))

    # Clear existing data rows (row 4+) up to summary
    for row in range(4, ws_overview.max_row + 1):
        for col in range(1, 11):
            cell = ws_overview.cell(row=row, column=col)
            cell.value = None
            cell.font = BODY_FONT
            cell.fill = PatternFill()

    # Build overview rows
    overview_rows = []
    for rid, group in cr_groups.items():
        teams_in_cr = sorted(set(t["team"] for t in group))
        team_count = len(teams_in_cr)

        # Use first task for title/dept, highest priority
        first = group[0]
        title = clean_title(first["task_name"])
        dept = first["dept_tag"]
        priority = first["priority"]

        # Per-team status (pick worst status if team appears multiple times)
        STATUS_RANK = {"Overdue": 0, "Blocked": 1, "Doing": 2, "UAT": 3, "Todo": 4, "Done": 5, "Rejected": 6}

        def team_status(team_name):
            team_tasks = [t for t in group if t["team"] == team_name]
            if not team_tasks:
                return "\u2014"
            # Pick worst status
            return min(team_tasks, key=lambda t: STATUS_RANK.get(t["status"], 99))["status"]

        mobile_status = team_status("Team Mobile")
        devb_status   = team_status("Dev Team B")
        mw_status     = team_status("MW Team")

        per_team = [mobile_status, devb_status, mw_status]
        overall = derive_overall_status(per_team)

        # Build notes from per-team context
        notes_parts = []
        for t in group:
            if t["status"] == "Overdue" and t["deadline"]:
                notes_parts.append(f"{t['team']}: overdue (deadline {t['deadline'].strftime('%d/%m/%Y')})")
            elif t["status"] == "Overdue":
                notes_parts.append(f"{t['team']}: overdue")
            elif not t["deadline"] and t["status"] == "Doing":
                notes_parts.append(f"{t['team']}: no deadline set")
        notes = ". ".join(notes_parts) if notes_parts else ""

        teams_label = f"{team_count} team{'s' if team_count > 1 else ''}"

        overview_rows.append([
            rid, title, dept, priority, teams_label,
            mobile_status, devb_status, mw_status,
            overall, notes,
        ])

    # Sort: multi-team first, then by priority, then by request_id
    PRIORITY_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    overview_rows.sort(key=lambda r: (
        -int(r[4].split()[0]),                    # team count descending
        PRIORITY_RANK.get(r[3], 99),              # priority
        r[0],                                      # request_id
    ))

    # Write overview rows
    for i, row_data in enumerate(overview_rows, start=4):
        for j, val in enumerate(row_data):
            cell = ws_overview.cell(row=i, column=j + 1, value=val)
            cell.alignment = LEFT if j in [1, 9] else CENTER
            cell.border = BORDER

            # Color status columns (F=5, G=6, H=7, I=8 → j=5,6,7,8)
            if j in [5, 6, 7, 8]:
                cell.font = STATUS_FONT.get(val, BODY_FONT)
                fill = STATUS_FILL.get(val)
                if fill:
                    cell.fill = fill
            elif j == 3:  # Priority
                cell.font = BOLD_FONT
                fill = PRIORITY_FILL.get(val)
                if fill:
                    cell.fill = fill
            else:
                cell.font = BODY_FONT

        ws_overview.row_dimensions[i].height = 26

    # Summary section
    last_data_row = 4 + len(overview_rows) - 1
    summary_start = last_data_row + 2

    ws_overview.merge_cells(f"A{summary_start}:J{summary_start}")
    ws_overview.cell(row=summary_start, column=1, value="Summary").font = Font(
        name="Segoe UI", size=12, bold=True, color="1B2A4A"
    )

    summary_formulas = [
        ("Total unique CRs:",          f'=COUNTA(A4:A{last_data_row})'),
        ("Multi-team CRs:",            f'=COUNTIF(E4:E{last_data_row},"2 teams")+COUNTIF(E4:E{last_data_row},"3 teams")'),
        ("CRs with Overdue status:",   f'=COUNTIF(I4:I{last_data_row},"Overdue")'),
        ("CRs fully Done:",            f'=COUNTIF(I4:I{last_data_row},"Done")'),
        ("CRs Partially Done:",        f'=COUNTIF(I4:I{last_data_row},"Partially Done")'),
        ("CRs In Progress:",           f'=COUNTIF(I4:I{last_data_row},"In Progress")'),
        ("CRs in UAT:",                f'=COUNTIF(I4:I{last_data_row},"UAT")'),
        ("CRs Blocked:",               f'=COUNTIF(I4:I{last_data_row},"Blocked")'),
        ("CRs Rejected:",              f'=COUNTIF(I4:I{last_data_row},"Rejected")'),
    ]

    for si, (label, formula) in enumerate(summary_formulas, summary_start + 1):
        ws_overview.cell(row=si, column=1, value=label).font = BOLD_FONT
        ws_overview.cell(row=si, column=1).alignment = LEFT
        cell = ws_overview.cell(row=si, column=2, value=formula)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="1B2A4A")
        cell.alignment = CENTER

    overview_count = len(overview_rows)
    multi_count = sum(1 for r in overview_rows if r[4] != "1 team")
    print(f"CR Overview: {overview_count} CRs ({multi_count} multi-team)")

    # ── Update Dashboard Summary formulas ────────────────────────────
    # The COUNTIF layout only makes sense on the template's pre-built
    # sheet — skip when the template doesn't include it.
    if "Dashboard Summary" in wb.sheetnames:
        ws_dash = wb["Dashboard Summary"]
        last_tracker_row = tracker_count + 2  # header is row 2, data starts row 3

        # teams_list comes from the template's Team Config sheet
        # Order matches the template header row (Todo..Blocked); UAT and Rejected
        # are appended with their own headers since the template predates them.
        statuses = ["Todo", "Doing", "Done", "Overdue", "Blocked", "UAT", "Rejected"]
        phases = ["Q1", "Q2", "Q3", "Q4", "Backlog"]

        for extra_col, extra_status in [(8, "UAT"), (9, "Rejected")]:
            hdr = ws_dash.cell(row=3, column=extra_col, value=extra_status)
            hdr.font = HEADER_FONT
            hdr.fill = HEADER_FILL
            hdr.alignment = CENTER

        for idx, team in enumerate(teams_list, 4):
            ws_dash.cell(row=idx, column=2).value = f"=COUNTIF('Task Tracker'!D:D,\"{team}\")"
            for si, status in enumerate(statuses, 3):
                ws_dash.cell(row=idx, column=si).value = (
                    f"=COUNTIFS('Task Tracker'!D:D,\"{team}\",'Task Tracker'!H:H,\"{status}\")"
                )

        # Phase column moved from I to J after the Stage column was inserted
        for idx, team in enumerate(teams_list, 11):
            for pi, phase in enumerate(phases, 2):
                ws_dash.cell(row=idx, column=pi).value = (
                    f"=COUNTIFS('Task Tracker'!D:D,\"{team}\",'Task Tracker'!J:J,\"{phase}\")"
                )
    else:
        print("Dashboard Summary: sheet not in template — skipped")

    # ── Star schema + snapshot history ───────────────────────────────
    write_star_schema(wb, tasks, TEAM_LEADS, report_date)
    csv_path, old_n, new_n = append_snapshot_csv(tasks, report_date, output_path)
    print(f"Star schema: fact_task_snapshot + dims + bridges written")
    print(f"Snapshot CSV: {csv_path} ({old_n} historical rows kept, {new_n} added for {report_date})")

    js_path = write_dashboard_js(tasks, TEAM_LEADS, report_date, output_path)
    print(f"Dashboard data: {js_path}")

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\nOutput saved: {output_path}")
    print("Sheets updated: Task Tracker, CR Overview, Dashboard Summary, star schema")


if __name__ == "__main__":
    main()
