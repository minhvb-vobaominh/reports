"""
Convert Base Wework tasks export → task tracker with CR Overview.

Populates both the "Task Tracker" and "CR Overview" sheets.
Auto-detects column layout from the header row (supports both 7-col and 20-col exports).

Usage:
    python convert_base_to_tracker.py [source.xls] [template.xlsx] [output.xlsx]

Defaults:
    source   = task-input.xls
    template = task-template.xlsx
    output   = task-output.xlsx
"""

import sys
import io
import os
import re
import argparse
from collections import OrderedDict
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CLI ──────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description="Convert Base Wework export → task tracker")
    p.add_argument("source", nargs="?", default="task-input.xls", help="Source .xls from Base Wework")
    p.add_argument("template", nargs="?", default="task-template.xlsx", help="Template .xlsx")
    p.add_argument("output", nargs="?", default="task-output.xlsx", help="Output .xlsx")
    return p.parse_args()


# ── Styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
BODY_FONT   = Font(name="Segoe UI", size=10, color="333333")
BOLD_FONT   = Font(name="Segoe UI", size=10, bold=True, color="333333")
MUTED_FONT  = Font(name="Segoe UI", size=9, color="64748B", italic=True)
BORDER = Border(bottom=Side(style="thin", color="D0D9E8"),
                right=Side(style="thin", color="D0D9E8"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

STATUS_FILL = {
    "Done":           PatternFill("solid", fgColor="DCFCE7"),
    "Doing":          PatternFill("solid", fgColor="DBEAFE"),
    "Overdue":        PatternFill("solid", fgColor="FEE2E2"),
    "Blocked":        PatternFill("solid", fgColor="FEF3C7"),
    "Todo":           PatternFill("solid", fgColor="F1F5F9"),
    "Not Started":    PatternFill("solid", fgColor="F1F5F9"),
    "In Progress":    PatternFill("solid", fgColor="DBEAFE"),
    "Partially Done": PatternFill("solid", fgColor="FEF9C3"),
}
STATUS_FONT = {
    "Done":           Font(name="Segoe UI", size=10, bold=True, color="15803D"),
    "Doing":          Font(name="Segoe UI", size=10, bold=True, color="1D4ED8"),
    "Overdue":        Font(name="Segoe UI", size=10, bold=True, color="B91C1C"),
    "Blocked":        Font(name="Segoe UI", size=10, bold=True, color="92400E"),
    "Todo":           Font(name="Segoe UI", size=10, color="475569"),
    "Not Started":    Font(name="Segoe UI", size=10, color="475569"),
    "In Progress":    Font(name="Segoe UI", size=10, bold=True, color="1D4ED8"),
    "Partially Done": Font(name="Segoe UI", size=10, bold=True, color="854D0E"),
    "\u2014":         Font(name="Segoe UI", size=10, color="CBD5E1"),
}
PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="FEE2E2"),
    "High":     PatternFill("solid", fgColor="FEF3C7"),
    "Medium":   PatternFill("solid", fgColor="DBEAFE"),
    "Low":      PatternFill("solid", fgColor="F1F5F9"),
}


# ── Team lead mapping ───────────────────────────────────────────────
LEAD_MAP = {
    "phongvo2440":    ("Team Mobile", "phongvo@phs.vn"),
    "minhvo2565":     ("Dev Team B",  "minhvo@phs.vn"),
    "huunguyen2525":  ("MW Team",     "huunguyen@phs.vn"),
}


# ── Status mapping (Vietnamese → English) ───────────────────────────
STATUS_MAP = {
    "Hoàn thành":      "Done",
    "Hoàn thành muộn": "Done",   # completed late → still Done
    "Đang làm":        "Doing",
    "Quá hạn":         "Overdue",
    "Chưa bắt đầu":   "Todo",
}

# ── Base Wework column indices ──────────────────────────────────────
# Auto-detected from header row; these are the defaults for the 20-col export.
COL_DEFAULTS = {
    "task_name":    0,   # Tên công việc
    "assigner":     1,   # Người giao việc
    "executor":     2,   # Người thực hiện
    "watcher":      3,   # Người theo dõi
    "urgent":       4,   # Khẩn cấp
    "important":    5,   # Quan trọng
    "labels":       6,   # Danh sách nhãn
    "start_date":   7,   # Ngày bắt đầu
    "deadline":     8,   # Thời hạn
    "completed":    9,   # Hoàn thành thực tế
    "description":  10,  # Mô tả công việc
    "status":       11,  # Trạng thái
    "result":       12,  # Kết quả công việc
    "objective":    13,  # Mục tiêu
    "created":      14,  # Ngày tạo
    "task_id":      15,  # Mã công việc (ID)
    "parent_id":    16,  # Mã công việc cha (ID)
    "metatype":     17,  # Metatype
    "executor_name": 18, # Họ và tên người nhận việc
    "assigner_name": 19, # Họ và tên người giao việc
}

# Known header names → column keys (Vietnamese)
HEADER_MAP = {
    "Tên công việc":              "task_name",
    "Người giao việc":            "assigner",
    "Người thực hiện":            "executor",
    "Người theo dõi":             "watcher",
    "Khẩn cấp":                   "urgent",
    "Quan trọng":                  "important",
    "Danh sách nhãn":             "labels",
    "Ngày bắt đầu":              "start_date",
    "Thời hạn":                    "deadline",
    "Hoàn thành thực tế":         "completed",
    "Mô tả công việc":           "description",
    "Trạng thái":                  "status",
    "Kết quả công việc":          "result",
    "Mục tiêu":                    "objective",
    "Ngày tạo":                    "created",
    "Mã công việc (ID)":          "task_id",
    "Mã công việc cha (ID)":      "parent_id",
    "Metatype":                    "metatype",
    "Họ và tên người nhận việc":  "executor_name",
    "Họ và tên người giao việc":  "assigner_name",
}


def detect_columns(sheet):
    """Auto-detect column indices from the header row (row 0).
    Falls back to COL_DEFAULTS if headers don't match."""
    cols = dict(COL_DEFAULTS)
    header_row = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    matched = 0
    for col_idx, header in enumerate(header_row):
        if header in HEADER_MAP:
            cols[HEADER_MAP[header]] = col_idx
            matched += 1
    return cols, matched


# ── Helpers ──────────────────────────────────────────────────────────
def parse_request_id(task_name):
    """Extract request-id from task name like [CR] [2275303] ... or [SDK] [2400512] ..."""
    m = re.search(r"\[(?:CR|SDK|INB)\]\s*\[(\d+)\]", task_name)
    if m:
        return m.group(1)
    # Try pattern like [CR][2372910] (no space)
    m = re.search(r"\[(?:CR|SDK|INB)\]\[(\d+)\]", task_name)
    if m:
        return m.group(1)
    return None


def parse_dept_tag(task_name):
    """Extract department tag like [RMD], [SS], [MKT] from task name."""
    # Skip the first bracket group ([CR]/[SDK]) and request-id, find dept tag
    m = re.search(r"\[\d+\]\s*\[([A-Z]{2,4})\]", task_name)
    if m:
        return m.group(1)
    return ""


def parse_category(task_name):
    """Determine category from task name prefix."""
    if task_name.startswith("[SDK]"):
        return "SDK"
    if task_name.startswith("[CR]") or task_name.startswith("[CR]["):
        return "CR"
    if task_name.startswith("[MKT]") or task_name.startswith("[INB]"):
        return "Internal"
    return "CR"


def parse_date_cell(value, datemode):
    """Parse date from xlrd cell value — handles both string and float (serial) dates."""
    if isinstance(value, float) and value > 0:
        try:
            dt = xlrd.xldate_as_datetime(value, datemode)
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return ""
    if isinstance(value, str) and value.strip():
        return value.strip()
    return ""


def clean_title(task_name):
    """Extract a clean CR title from the full task name.
    Removes [CR], [request-id], [dept], [No.XX-YYYY] prefixes."""
    title = task_name
    # Remove leading [CR] / [SDK] / [INB]
    title = re.sub(r"^\[(?:CR|SDK|INB)\]\s*", "", title)
    # Remove [request-id]
    title = re.sub(r"^\[?\d+\]?\s*", "", title)
    # Remove [dept]
    title = re.sub(r"^\[[A-Z]{2,4}\]\s*", "", title)
    # Remove [No.XX-YYYY]
    title = re.sub(r"^\[No\.\d+-\d+\]\s*", "", title)
    return title.strip()


def detect_section(section_name):
    """Parse section header to get team and phase.
    e.g. 'Team Mobile - Q2:' → ('Team Mobile', 'Q2')
         'Dev Team B - Backlog:' → ('Dev Team B', 'Backlog')
         'CRs in Coding:' → (None, 'Q2')  # cross-team, use lead to determine
    """
    section_name = section_name.rstrip(":")
    if "CRs in Coding" in section_name:
        return None, "Q2"

    for team in ["Team Mobile", "Dev Team B", "MW Team"]:
        if team in section_name:
            if "Backlog" in section_name:
                return team, "Backlog"
            m = re.search(r"Q(\d)", section_name)
            if m:
                return team, f"Q{m.group(1)}"
            return team, "Q2"
    return None, "Q2"


def derive_overall_status(statuses):
    """Derive overall CR status from per-team statuses."""
    active = [s for s in statuses if s != "\u2014"]
    if not active:
        return "Not Started"
    if any(s == "Overdue" for s in active):
        return "Overdue"
    if any(s == "Blocked" for s in active):
        return "Blocked"
    if all(s == "Done" for s in active):
        return "Done"
    if all(s == "Todo" for s in active):
        return "Not Started"
    if any(s == "Done" for s in active) and any(s != "Done" for s in active):
        return "Partially Done"
    return "In Progress"


def derive_priority(statuses_with_deadlines):
    """Simple priority heuristic based on overdue count and deadline proximity."""
    overdue_count = sum(1 for s, _ in statuses_with_deadlines if s == "Overdue")
    if overdue_count >= 2:
        return "Critical"
    if overdue_count == 1:
        return "High"
    has_deadline = any(d for _, d in statuses_with_deadlines)
    if has_deadline:
        return "Medium"
    return "Low"


# ── Main ─────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    source_path   = os.path.join(script_dir, args.source)   if not os.path.isabs(args.source)   else args.source
    template_path = os.path.join(script_dir, args.template)  if not os.path.isabs(args.template) else args.template
    output_path   = os.path.join(script_dir, args.output)    if not os.path.isabs(args.output)   else args.output

    # ── Read source .xls ─────────────────────────────────────────────
    src_wb = xlrd.open_workbook(source_path)
    src_ws = src_wb.sheet_by_index(0)
    datemode = src_wb.datemode

    print(f"Source: {source_path} ({src_ws.nrows} rows, {src_ws.ncols} cols)")

    # Auto-detect column layout from header row
    C, matched = detect_columns(src_ws)
    print(f"Column detection: {matched}/{len(HEADER_MAP)} headers matched")

    # Parse all task rows, skipping headers and section dividers
    tasks = []
    current_section_team = None
    current_phase = "Q2"

    def cell_str(row, key):
        return str(src_ws.cell_value(row, C[key])).strip()

    def cell_raw(row, key):
        return src_ws.cell_value(row, C[key])

    for row_idx in range(1, src_ws.nrows):
        task_name  = cell_str(row_idx, "task_name")
        assigner   = cell_str(row_idx, "assigner")
        deadline   = cell_raw(row_idx, "deadline")
        completed  = cell_raw(row_idx, "completed")
        status_vn  = cell_str(row_idx, "status")
        created    = cell_str(row_idx, "created")
        task_id    = cell_str(row_idx, "task_id")
        executor   = cell_str(row_idx, "executor") if "executor" in C else ""
        executor_name = cell_str(row_idx, "executor_name") if "executor_name" in C else ""

        # Section header row (no assigner, name ends with ':')
        if not assigner and ":" in task_name:
            current_section_team, current_phase = detect_section(task_name)
            continue

        # Skip empty rows
        if not task_name:
            continue

        # Determine team
        if current_section_team:
            team = current_section_team
        elif assigner in LEAD_MAP:
            team = LEAD_MAP[assigner][0]
        else:
            team = "Unknown"

        # Get lead email
        lead_email = LEAD_MAP.get(assigner, (None, assigner))[1]

        # Parse fields
        request_id  = parse_request_id(task_name)
        dept_tag    = parse_dept_tag(task_name)
        category    = parse_category(task_name)
        status      = STATUS_MAP.get(status_vn, status_vn)
        deadline_str  = parse_date_cell(deadline, datemode)
        completed_str = parse_date_cell(completed, datemode)

        # Backlog tasks → Todo
        phase = current_phase
        if phase == "Backlog" and status == "Doing":
            status = "Todo"

        # Clean up task_id (remove .0 from float)
        if task_id.endswith(".0"):
            task_id = task_id[:-2]

        # Determine assignee: prefer executor_name (full name), fall back to executor (username)
        assignee = executor_name if executor_name else executor

        tasks.append({
            "request_id": request_id or "",
            "task_id":    task_id,
            "task_name":  task_name,
            "team":       team,
            "category":   category,
            "dept_tag":   dept_tag,
            "priority":   "",  # will be set later
            "status":     status,
            "phase":      phase,
            "assignee":   assignee,
            "lead_email": lead_email,
            "created":    created,
            "deadline":   deadline_str,
            "completed":  completed_str,
            "dependencies": "",
            "notes":      "",
        })

    print(f"Parsed: {len(tasks)} tasks")

    # ── Build CR groups ──────────────────────────────────────────────
    # Group by request_id to detect multi-team CRs and derive overview
    cr_groups = OrderedDict()  # request_id → list of task dicts
    no_id_tasks = []

    for t in tasks:
        rid = t["request_id"]
        if rid:
            cr_groups.setdefault(rid, []).append(t)
        else:
            no_id_tasks.append(t)

    # For multi-team CRs, set dependencies and priority
    for rid, group in cr_groups.items():
        teams_in_cr = set(t["team"] for t in group)
        is_multi = len(teams_in_cr) > 1

        # Priority heuristic
        statuses_deadlines = [(t["status"], t["deadline"]) for t in group]
        priority = derive_priority(statuses_deadlines)

        for t in group:
            t["priority"] = priority
            if is_multi:
                t["dependencies"] = rid  # link to same CR

    # Single-team CRs without priority yet
    for t in tasks:
        if not t["priority"]:
            if t["status"] == "Overdue":
                t["priority"] = "High"
            elif t["phase"] == "Backlog":
                t["priority"] = "Low"
            else:
                t["priority"] = "Medium"

    # ── Load template ────────────────────────────────────────────────
    wb = load_workbook(template_path)

    # ── Populate Task Tracker ────────────────────────────────────────
    ws_tracker = wb["Task Tracker"]

    # Clear existing sample data (rows 3+)
    for row in range(3, ws_tracker.max_row + 1):
        for col in range(1, 17):
            ws_tracker.cell(row=row, column=col).value = None
            ws_tracker.cell(row=row, column=col).font = BODY_FONT
            ws_tracker.cell(row=row, column=col).fill = PatternFill()

    # Write tasks
    tracker_fields = [
        "request_id", "task_id", "task_name", "team", "category",
        "dept_tag", "priority", "status", "phase", "assignee",
        "lead_email", "created", "deadline", "completed",
        "dependencies", "notes",
    ]
    wide_cols = {2, 14, 15}  # 0-indexed: task_name, dependencies, notes

    for i, t in enumerate(tasks, start=3):
        for j, field in enumerate(tracker_fields):
            cell = ws_tracker.cell(row=i, column=j + 1, value=t[field])
            cell.font = BODY_FONT
            cell.alignment = LEFT if j in wide_cols else CENTER
            cell.border = BORDER
        ws_tracker.row_dimensions[i].height = 24

    tracker_count = len(tasks)
    print(f"Task Tracker: {tracker_count} rows written")

    # ── Populate CR Overview ─────────────────────────────────────────
    ws_overview = wb["CR Overview"]

    # Unmerge all merged cells in rows 4+ before clearing
    merges_to_remove = [m for m in ws_overview.merged_cells.ranges if m.min_row >= 4]
    for m in merges_to_remove:
        ws_overview.unmerge_cells(str(m))

    # Clear existing data rows (row 4+) up to summary
    for row in range(4, ws_overview.max_row + 1):
        for col in range(1, 11):
            cell = ws_overview.cell(row=row, column=col)
            cell.value = None
            cell.font = BODY_FONT
            cell.fill = PatternFill()

    # Build overview rows
    overview_rows = []
    for rid, group in cr_groups.items():
        teams_in_cr = sorted(set(t["team"] for t in group))
        team_count = len(teams_in_cr)

        # Use first task for title/dept, highest priority
        first = group[0]
        title = clean_title(first["task_name"])
        dept = first["dept_tag"]
        priority = first["priority"]

        # Per-team status (pick worst status if team appears multiple times)
        STATUS_RANK = {"Overdue": 0, "Blocked": 1, "Doing": 2, "Todo": 3, "Done": 4}

        def team_status(team_name):
            team_tasks = [t for t in group if t["team"] == team_name]
            if not team_tasks:
                return "\u2014"
            # Pick worst status
            return min(team_tasks, key=lambda t: STATUS_RANK.get(t["status"], 99))["status"]

        mobile_status = team_status("Team Mobile")
        devb_status   = team_status("Dev Team B")
        mw_status     = team_status("MW Team")

        per_team = [mobile_status, devb_status, mw_status]
        overall = derive_overall_status(per_team)

        # Build notes from per-team context
        notes_parts = []
        for t in group:
            if t["status"] == "Overdue" and t["deadline"]:
                notes_parts.append(f"{t['team']}: overdue (deadline {t['deadline'].split(' ')[0]})")
            elif t["status"] == "Overdue":
                notes_parts.append(f"{t['team']}: overdue")
            elif not t["deadline"] and t["status"] == "Doing":
                notes_parts.append(f"{t['team']}: no deadline set")
        notes = ". ".join(notes_parts) if notes_parts else ""

        teams_label = f"{team_count} team{'s' if team_count > 1 else ''}"

        overview_rows.append([
            rid, title, dept, priority, teams_label,
            mobile_status, devb_status, mw_status,
            overall, notes,
        ])

    # Sort: multi-team first, then by priority, then by request_id
    PRIORITY_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    overview_rows.sort(key=lambda r: (
        -int(r[4].split()[0]),                    # team count descending
        PRIORITY_RANK.get(r[3], 99),              # priority
        r[0],                                      # request_id
    ))

    # Write overview rows
    for i, row_data in enumerate(overview_rows, start=4):
        for j, val in enumerate(row_data):
            cell = ws_overview.cell(row=i, column=j + 1, value=val)
            cell.alignment = LEFT if j in [1, 9] else CENTER
            cell.border = BORDER

            # Color status columns (F=5, G=6, H=7, I=8 → j=5,6,7,8)
            if j in [5, 6, 7, 8]:
                cell.font = STATUS_FONT.get(val, BODY_FONT)
                fill = STATUS_FILL.get(val)
                if fill:
                    cell.fill = fill
            elif j == 3:  # Priority
                cell.font = BOLD_FONT
                fill = PRIORITY_FILL.get(val)
                if fill:
                    cell.fill = fill
            else:
                cell.font = BODY_FONT

        ws_overview.row_dimensions[i].height = 26

    # Summary section
    last_data_row = 4 + len(overview_rows) - 1
    summary_start = last_data_row + 2

    ws_overview.merge_cells(f"A{summary_start}:J{summary_start}")
    ws_overview.cell(row=summary_start, column=1, value="Summary").font = Font(
        name="Segoe UI", size=12, bold=True, color="1B2A4A"
    )

    summary_formulas = [
        ("Total unique CRs:",          f'=COUNTA(A4:A{last_data_row})'),
        ("Multi-team CRs:",            f'=COUNTIF(E4:E{last_data_row},"2 teams")+COUNTIF(E4:E{last_data_row},"3 teams")'),
        ("CRs with Overdue status:",   f'=COUNTIF(I4:I{last_data_row},"Overdue")'),
        ("CRs fully Done:",            f'=COUNTIF(I4:I{last_data_row},"Done")'),
        ("CRs Partially Done:",        f'=COUNTIF(I4:I{last_data_row},"Partially Done")'),
        ("CRs In Progress:",           f'=COUNTIF(I4:I{last_data_row},"In Progress")'),
        ("CRs Blocked:",               f'=COUNTIF(I4:I{last_data_row},"Blocked")'),
    ]

    for si, (label, formula) in enumerate(summary_formulas, summary_start + 1):
        ws_overview.cell(row=si, column=1, value=label).font = BOLD_FONT
        ws_overview.cell(row=si, column=1).alignment = LEFT
        cell = ws_overview.cell(row=si, column=2, value=formula)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="1B2A4A")
        cell.alignment = CENTER

    overview_count = len(overview_rows)
    multi_count = sum(1 for r in overview_rows if r[4] != "1 team")
    print(f"CR Overview: {overview_count} CRs ({multi_count} multi-team)")

    # ── Update Dashboard Summary formulas ────────────────────────────
    ws_dash = wb["Dashboard Summary"]
    last_tracker_row = tracker_count + 2  # header is row 2, data starts row 3

    teams_list = ["Team Mobile", "Dev Team B", "MW Team"]
    statuses = ["Todo", "Doing", "Done", "Overdue", "Blocked"]
    phases = ["Q1", "Q2", "Q3", "Q4", "Backlog"]

    for idx, team in enumerate(teams_list, 4):
        ws_dash.cell(row=idx, column=2).value = f"=COUNTIF('Task Tracker'!D:D,\"{team}\")"
        for si, status in enumerate(statuses, 3):
            ws_dash.cell(row=idx, column=si).value = (
                f"=COUNTIFS('Task Tracker'!D:D,\"{team}\",'Task Tracker'!H:H,\"{status}\")"
            )

    for idx, team in enumerate(teams_list, 11):
        for pi, phase in enumerate(phases, 2):
            ws_dash.cell(row=idx, column=pi).value = (
                f"=COUNTIFS('Task Tracker'!D:D,\"{team}\",'Task Tracker'!I:I,\"{phase}\")"
            )

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\nOutput saved: {output_path}")
    print("Sheets updated: Task Tracker, CR Overview, Dashboard Summary")


if __name__ == "__main__":
    main()
